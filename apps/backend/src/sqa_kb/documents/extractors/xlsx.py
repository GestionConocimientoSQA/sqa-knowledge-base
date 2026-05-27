"""XlsxExtractor — lee texto de un `.xlsx` (Fase 4.3).

Usa openpyxl en modo `read_only` + `data_only` (lee valores cacheados,
no fórmulas). Cada hoja es una `ExtractedSection` (`title` = nombre de
hoja, `content` = filas serializadas con ' | ' entre celdas).
"""

from __future__ import annotations

import io

from openpyxl import load_workbook

from sqa_kb.documents.extractors.base import ExtractedDocument, ExtractedSection

EXTENSIONS = ("xlsx",)


class XlsxExtractor:
    """Implementa `DocumentExtractor` para `.xlsx`."""

    @property
    def extensions(self) -> tuple[str, ...]:
        return EXTENSIONS

    def extract(self, data: bytes) -> ExtractedDocument:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        sections: list[ExtractedSection] = []
        all_text: list[str] = []

        try:
            for ws in wb.worksheets:
                rows_text: list[str] = []
                for row in ws.iter_rows(values_only=True):
                    cells = [str(c) for c in row if c is not None and str(c).strip()]
                    if cells:
                        rows_text.append(" | ".join(cells))
                if rows_text:
                    body = "\n".join(rows_text)
                    sections.append(ExtractedSection(title=ws.title, content=body))
                    all_text.append(f"[{ws.title}]\n{body}")
        finally:
            wb.close()

        return ExtractedDocument(
            text="\n\n".join(all_text).strip(),
            sections=tuple(sections),
            page_count=len(sections),
        )

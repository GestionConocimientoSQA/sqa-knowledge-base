"""XlsxGenerator — planillas Excel con branding SQA (Fase 4.2).

Pensado para el tipo FORM (formatos/plantillas tabulares), pero acepta
cualquier `DocumentContent`.

Estructura del workbook:
- Hoja "Metadata": tabla de propiedades del documento (carpeta, tipo,
  versión, fecha, autor, anonimizado).
- Hoja "Contenido": el topic + cada body_block en filas.
- Hoja "Precisiones" (si hay qa_pairs): pregunta | respuesta.

Branding: header de tablas con fondo azul corp + texto blanco, fuente
del cuerpo, freeze de la fila header.
"""

from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from sqa_kb.documents import branding
from sqa_kb.documents.doc_types import category_label, doc_type_label
from sqa_kb.documents.generators.base import GeneratedFile
from sqa_kb.documents.models import DocumentContent

MEDIA_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
EXTENSION = "xlsx"

_HEADER_FILL = PatternFill(
    start_color=branding.COLOR_FONDO_TABLA_HEADER.hex,
    end_color=branding.COLOR_FONDO_TABLA_HEADER.hex,
    fill_type="solid",
)
_HEADER_FONT = Font(
    name=branding.FONT_BODY,
    bold=True,
    color=branding.BLANCO.hex,
    size=branding.SIZE_BODY,
)
_BODY_FONT = Font(
    name=branding.FONT_BODY,
    color=branding.COLOR_CUERPO.hex,
    size=branding.SIZE_BODY,
)


class XlsxGenerator:
    """Genera `.xlsx` con branding SQA. Implementa `DocumentGenerator`."""

    @property
    def extension(self) -> str:
        return EXTENSION

    @property
    def media_type(self) -> str:
        return MEDIA_TYPE

    def generate(self, content: DocumentContent) -> GeneratedFile:
        wb = Workbook()
        # La primera hoja viene por default — la usamos como Metadata.
        meta_ws = wb.active
        meta_ws.title = "Metadata"
        self._fill_metadata(meta_ws, content)

        content_ws = wb.create_sheet("Contenido")
        self._fill_content(content_ws, content)

        if content.qa_pairs:
            qa_ws = wb.create_sheet("Precisiones")
            self._fill_qa(qa_ws, content)

        buffer = io.BytesIO()
        wb.save(buffer)
        return GeneratedFile(
            filename=f"{content.document_id}.{EXTENSION}",
            media_type=MEDIA_TYPE,
            data=buffer.getvalue(),
        )

    # -- hojas --------------------------------------------------------------

    def _fill_metadata(self, ws: Worksheet, content: DocumentContent) -> None:
        self._write_header(ws, ["Propiedad", "Valor"])
        autor = content.author_name
        if content.author_role:
            autor = f"{content.author_name} ({content.author_role})"
        rows = [
            ("Título", content.title),
            ("Carpeta", f"{content.category} — {category_label(content.category)}"),
            ("Tipo", f"{content.document_type} — {doc_type_label(content.document_type)}"),
            ("Versión", content.version),
            ("Fecha", content.fecha.strftime("%Y-%m-%d")),
            ("Autor", autor),
            ("Anonimizado", "Sí" if content.is_anonymized else "No"),
        ]
        for prop, value in rows:
            ws.append([prop, value])
        self._style_body(ws, start_row=2, cols=2)
        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 60

    def _fill_content(self, ws: Worksheet, content: DocumentContent) -> None:
        self._write_header(ws, ["#", "Contenido"])
        ws.append([0, content.topic])  # fila 2: el tema
        for i, block in enumerate(content.body_blocks, start=1):
            ws.append([i, block])
        self._style_body(ws, start_row=2, cols=2)
        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 90

    def _fill_qa(self, ws: Worksheet, content: DocumentContent) -> None:
        self._write_header(ws, ["Pregunta", "Respuesta"])
        for qa in content.qa_pairs:
            ws.append([qa.question, qa.answer])
        self._style_body(ws, start_row=2, cols=2)
        ws.column_dimensions["A"].width = 45
        ws.column_dimensions["B"].width = 70

    # -- estilos ------------------------------------------------------------

    def _write_header(self, ws: Worksheet, headers: list[str]) -> None:
        ws.append(headers)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = Alignment(vertical="center")
        ws.freeze_panes = "A2"

    def _style_body(self, ws: Worksheet, *, start_row: int, cols: int) -> None:
        for row in ws.iter_rows(
            min_row=start_row, max_row=ws.max_row, min_col=1, max_col=cols
        ):
            for cell in row:
                cell.font = _BODY_FONT
                cell.alignment = Alignment(vertical="top", wrap_text=True)
